"""The XLSX mentorship workbook.

Nine sheets, all driven off one Excel Table so every number is auditable back
to the trade that produced it.

On pivot tables: openpyxl cannot emit a native PivotTable - the format is a
separate cache structure it does not write. Rather than ship a workbook that
needs a manual refresh, the pivot sheets are built as live formula grids
(SUMIFS/COUNTIFS over the Trades table) wired to dropdown slicers. They
recalculate the instant a filter changes, survive a round trip through Excel,
Google Sheets and LibreOffice, and - unlike a cached pivot - can never show
stale figures.

Design rules followed throughout:

  * Every derived cell is a formula, not a baked value, wherever Excel can do
    the arithmetic. A mentorship file the trader cannot interrogate teaches
    nothing.
  * Durations follow the brief: seconds under a minute, minutes under an hour,
    hours and minutes above that.
  * A statistic computed from too small a sample is labelled as such rather
    than quietly presented next to one computed from hundreds.
"""

from __future__ import annotations

import datetime as dt
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import absolute_coordinate, get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import db, metrics, model, paths, settings, stats

# -- palette ---------------------------------------------------------------
# Dark navy chrome, light body. Reads as modern on screen and still prints.
INK = "0F172A"
INK_SOFT = "1E293B"
SLATE = "475569"
MUTED = "94A3B8"
LINE = "E2E8F0"
CARD = "F8FAFC"
WHITE = "FFFFFF"
TEAL = "0D9488"
TEAL_SOFT = "CCFBF1"
RED = "DC2626"
RED_SOFT = "FEE2E2"
BLUE = "2563EB"
BLUE_SOFT = "DBEAFE"
AMBER = "D97706"
AMBER_SOFT = "FEF3C7"
VIOLET = "7C3AED"

# Charts need their categories and values to live in cells somewhere. Those
# helper tables sit out past the printed area and their columns are hidden, so
# the dashboard reads as a dashboard rather than as a chart plus its workings.
HELPER_COL = 22  # column V

MONEY = '#,##0.00;[Red]-#,##0.00'
MONEY_SIGN = '+#,##0.00;[Red]-#,##0.00;0.00'
PCT = '0.0"%"'
RATIO = '0.00'
INT = '#,##0'
DATE_TIME = 'yyyy-mm-dd hh:mm:ss'


def _fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def _thin(colour: str = LINE) -> Border:
    side = Side(style="thin", color=colour)
    return Border(left=side, right=side, top=side, bottom=side)


def register_styles(book: Workbook) -> None:
    """Named styles, registered once and reused - keeps the file small."""
    definitions = [
        ("ps_title", Font(name="Segoe UI Semibold", size=20, color=WHITE),
         _fill(INK), Alignment("left", "center")),
        ("ps_subtitle", Font(name="Segoe UI", size=10, color=MUTED),
         _fill(INK), Alignment("left", "center")),
        ("ps_section", Font(name="Segoe UI Semibold", size=12, color=INK),
         None, Alignment("left", "center")),
        ("ps_header", Font(name="Segoe UI Semibold", size=10, color=WHITE),
         _fill(INK_SOFT), Alignment("center", "center", wrap_text=True)),
        ("ps_label", Font(name="Segoe UI", size=9, color=SLATE),
         None, Alignment("left", "center")),
        ("ps_kpi_label", Font(name="Segoe UI", size=9, color=SLATE),
         _fill(CARD), Alignment("left", "center")),
        ("ps_kpi_value", Font(name="Segoe UI Semibold", size=16, color=INK),
         _fill(CARD), Alignment("left", "center")),
        ("ps_kpi_note", Font(name="Segoe UI", size=8, color=MUTED),
         _fill(CARD), Alignment("left", "center")),
        ("ps_cell", Font(name="Segoe UI", size=9, color=INK),
         None, Alignment("center", "center")),
        ("ps_cell_left", Font(name="Segoe UI", size=9, color=INK),
         None, Alignment("left", "center")),
        ("ps_note", Font(name="Segoe UI", size=9, color=SLATE, italic=True),
         None, Alignment("left", "top", wrap_text=True)),
        # Same look, no wrapping: lets a long note spill across the empty
        # columns to its right instead of being clipped to one cell.
        ("ps_note_flow", Font(name="Segoe UI", size=9, color=SLATE, italic=True),
         None, Alignment("left", "center")),
    ]
    for name, font, fill, alignment in definitions:
        if name in book.named_styles:
            continue
        style = NamedStyle(name=name)
        style.font = font
        if fill:
            style.fill = fill
        style.alignment = alignment
        book.add_named_style(style)


def _write(sheet, row: int, column: int, value, style: str | None = None,
           number_format: str | None = None, fill: str | None = None,
           font: Font | None = None, border: Border | None = None,
           align: Alignment | None = None):
    cell = sheet.cell(row=row, column=column, value=value)
    if style:
        cell.style = style
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = _fill(fill)
    if font:
        cell.font = font
    if border:
        cell.border = border
    if align:
        cell.alignment = align
    return cell


def _banner(sheet, title: str, subtitle: str, width: int = 12) -> None:
    """Dark title band across the top of a sheet."""
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    _write(sheet, 1, 1, title, "ps_title")
    _write(sheet, 2, 1, subtitle, "ps_subtitle")
    for column in range(1, width + 1):
        sheet.cell(row=1, column=column).fill = _fill(INK)
        sheet.cell(row=2, column=column).fill = _fill(INK)
    sheet.row_dimensions[1].height = 34
    sheet.row_dimensions[2].height = 18
    sheet.sheet_view.showGridLines = False


def _print_setup(book: Workbook, account: dict) -> None:
    """Make every sheet print like a report rather than a spreadsheet dump.

    Left to Excel's defaults the dashboard breaks across six sheets of paper
    with the KPI cards sliced down the middle. Fitting to one page wide is what
    makes a printed copy - or a PDF export, which is how these get shared with a
    mentor - readable.
    """
    footer_left = f"{paths.APP_TITLE} · account {account.get('login', '')}"
    for sheet in book.worksheets:
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_options.horizontalCentered = True
        sheet.oddFooter.left.text = footer_left
        sheet.oddFooter.left.size = 8
        sheet.oddFooter.left.color = SLATE
        sheet.oddFooter.right.text = "&P / &N"
        sheet.oddFooter.right.size = 8
        sheet.oddFooter.right.color = SLATE

        last = sheet.max_row
        # Print area stops short of the hidden helper columns: a print area that
        # includes them makes Excel reserve the width even though nothing shows.
        end_column = get_column_letter(min(sheet.max_column, HELPER_COL - 1))
        if last > 1:
            sheet.print_area = f"A1:{end_column}{last}"

    trades = book[SHEET_TRADES]
    # 59 columns will not fit one page wide legibly, so this one sheet prints
    # across as many pages as it needs with the header row repeated on each.
    trades.page_setup.fitToWidth = 0
    trades.print_title_rows = "1:1"


def _duration_seconds(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# -- the Trades sheet ------------------------------------------------------
# (header, key, source, number format, width). `source` is "p" for a position
# field and "m" for a computed metric. Order is deliberate: identity, then the
# plan, then the outcome, then the excursion forensics.
COLUMNS: list[tuple[str, str, str, str | None, int]] = [
    ("#", "trade_no", "m", INT, 6),
    ("Ticket", "ticket", "p", "0", 12),
    ("Symbol", "symbol", "p", None, 11),
    ("Side", "side", "p", None, 7),
    ("Volume", "volume", "p", "0.00", 8),
    ("Opened", "open_dt", "x", DATE_TIME, 19),
    ("Closed", "close_dt", "x", DATE_TIME, 19),
    ("Weekday", "weekday", "m", None, 10),
    ("Hour", "hour", "m", "0", 6),
    ("Session", "session", "m", None, 12),
    ("Week", "week", "m", None, 10),
    ("Month", "month", "m", None, 9),
    ("Duration", "duration_label", "m", None, 11),
    ("Duration (s)", "duration_s", "m", "0.0", 12),
    ("Duration (min)", "duration_minutes", "m", "0.00", 13),
    ("Duration band", "duration_bucket", "m", None, 13),
    ("Entry", "open_price", "p", None, 12),
    ("Exit", "close_price", "p", None, 12),
    ("SL initial", "sl_initial", "p", None, 12),
    ("TP initial", "tp_initial", "p", None, 12),
    ("SL final", "sl_final", "p", None, 12),
    ("SL moved", "sl_direction", "m", None, 11),
    ("SL move (pts)", "sl_move_points", "m", "0.0", 12),
    ("Risk", "risk_money", "m", MONEY, 10),
    ("Planned RR", "planned_rr", "m", RATIO, 11),
    ("R multiple", "r_multiple", "m", '+0.00;[Red]-0.00;0.00', 11),
    ("R band", "r_bucket", "m", None, 11),
    ("Gross", "gross_profit", "p", MONEY, 11),
    ("Costs", "costs", "m", MONEY, 10),
    ("Net", "net_profit", "p", MONEY_SIGN, 11),
    ("Outcome", "outcome", "m", None, 10),
    ("Exit reason", "exit_reason", "p", None, 14),
    ("Result (pts)", "result_points", "m", "0.0", 12),
    ("MFE", "mfe_money", "m", MONEY, 10),
    ("MAE", "mae_money", "m", MONEY, 10),
    ("MFE (R)", "mfe_r", "m", RATIO, 10),
    ("MAE (R)", "mae_r", "m", RATIO, 10),
    ("MFE price", "mfe_price", "m", None, 12),
    ("MAE price", "mae_price", "m", None, 12),
    ("Time to MFE", "time_to_mfe_label", "m", None, 12),
    ("Time to MAE", "time_to_mae_label", "m", None, 12),
    ("Heat %", "heat_pct", "m", PCT, 9),
    ("Capture", "capture_ratio", "m", RATIO, 9),
    ("Giveback", "giveback_money", "m", MONEY, 10),
    ("Entry eff.", "entry_efficiency", "m", RATIO, 10),
    ("Exit eff.", "exit_efficiency", "m", RATIO, 10),
    ("Total eff.", "total_efficiency", "m", RATIO, 10),
    ("MFE first", "mfe_before_mae", "m", None, 10),
    ("Near miss", "near_miss", "m", None, 10),
    ("Risk % bal.", "risk_pct_of_balance", "m", '0.00"%"', 11),
    ("Balance after", "balance_after", "m", MONEY, 13),
    ("Drawdown", "drawdown_money", "m", MONEY, 11),
    ("Drawdown %", "drawdown_pct", "m", PCT, 11),
    ("Streak", "streak", "m", '+0;-0;0', 8),
    ("Excursion src", "excursion_source", "m", None, 12),
    ("Samples", "excursion_samples", "m", INT, 10),
    ("Shots", "shot_count", "x", INT, 8),
    ("Folder", "folder", "x", None, 30),
    ("Note", "note", "p", None, 24),
]

TABLE_NAME = "Trades"
SHEET_TRADES = "Trades"


# Metrics stores these as 1/0. They read as Yes/No in a sheet a human is going
# to filter by hand, and a pivot row labelled "1" explains nothing.
_FLAG_HEADERS = {"MFE first", "Near miss"}


def _cell_value(position: dict, key: str, source: str, extras: dict,
                header: str = ""):
    if source == "p":
        value = position.get(key)
    elif source == "m":
        value = (position.get("metrics") or {}).get(key)
    else:
        value = extras.get(key)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if header in _FLAG_HEADERS and value is not None:
        return "Yes" if value else "No"
    return value


def build_trades(book: Workbook, positions: list[dict], shots: dict[int, list[dict]],
                 folders: dict[int, str]) -> tuple[str, int]:
    """The spine of the workbook: one row per closed position, as a real Table.

    Returns (table range, first data row). Everything else references this.
    """
    sheet = book.create_sheet(SHEET_TRADES)
    header_row = 1
    for index, (title, _key, _src, _fmt, width) in enumerate(COLUMNS, start=1):
        _write(sheet, header_row, index, title, "ps_header")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[header_row].height = 30

    for offset, position in enumerate(positions):
        row = header_row + 1 + offset
        extras = {
            "open_dt": model.broker_dt(position.get("open_time")),
            "close_dt": model.broker_dt(position.get("close_time")),
            "shot_count": len(shots.get(position["id"], [])),
            "folder": folders.get(position["id"], ""),
        }
        digits = int((position.get("metrics") or {}).get("digits") or 2)
        price_format = "0." + "0" * digits if digits else "0"
        for index, (title, key, source, number_format, _width) in enumerate(
            COLUMNS, start=1
        ):
            value = _cell_value(position, key, source, extras, title)
            cell = _write(sheet, row, index, value, "ps_cell")
            cell.number_format = number_format or price_format
            if number_format is None and not isinstance(value, (int, float)):
                cell.number_format = "General"
        sheet.row_dimensions[row].height = 16

    last_row = header_row + len(positions)
    last_column = get_column_letter(len(COLUMNS))
    ref = f"A{header_row}:{last_column}{max(last_row, header_row + 1)}"
    table = Table(displayName=TABLE_NAME, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)
    sheet.freeze_panes = "C2"
    sheet.sheet_view.showGridLines = False
    _decorate_trades(sheet, header_row + 1, last_row)
    return ref, header_row + 1


def col_of(header: str) -> str:
    """Column letter for a Trades header. Formulas elsewhere depend on this."""
    for index, (title, *_rest) in enumerate(COLUMNS, start=1):
        if title == header:
            return get_column_letter(index)
    raise KeyError(header)


def rng(header: str, first_row: int, last_row: int) -> str:
    """Absolute range for one Trades column, usable inside SUMIFS."""
    letter = col_of(header)
    return f"{SHEET_TRADES}!${letter}${first_row}:${letter}${last_row}"


def _decorate_trades(sheet, first_row: int, last_row: int) -> None:
    """Conditional formatting - the part that makes 73 rows readable at a glance."""
    if last_row < first_row:
        return

    def span(header: str) -> str:
        letter = col_of(header)
        return f"{letter}{first_row}:{letter}{last_row}"

    green_text = Font(color="065F46", bold=True)
    red_text = Font(color="991B1B", bold=True)

    for header in ("Net", "R multiple"):
        area = span(header)
        sheet.conditional_formatting.add(
            area,
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=_fill(TEAL_SOFT), font=green_text),
        )
        sheet.conditional_formatting.add(
            area,
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=_fill(RED_SOFT), font=red_text),
        )

    sheet.conditional_formatting.add(
        span("Outcome"),
        CellIsRule(operator="equal", formula=['"Win"'], fill=_fill(TEAL_SOFT)),
    )
    sheet.conditional_formatting.add(
        span("Outcome"),
        CellIsRule(operator="equal", formula=['"Loss"'], fill=_fill(RED_SOFT)),
    )

    # Heat above 100% means the trade went further against the trader than the
    # stop it was risking - the single most diagnostic cell in the sheet.
    sheet.conditional_formatting.add(
        span("Heat %"),
        CellIsRule(operator="greaterThanOrEqual", formula=["100"],
                   fill=_fill(RED_SOFT), font=red_text),
    )
    sheet.conditional_formatting.add(
        span("MFE"),
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color=TEAL, showValue=True),
    )
    sheet.conditional_formatting.add(
        span("MAE"),
        DataBarRule(start_type="min", end_type="num", end_value=0,
                    color=RED, showValue=True),
    )
    sheet.conditional_formatting.add(
        span("Capture"),
        ColorScaleRule(start_type="num", start_value=0, start_color="FECACA",
                       mid_type="num", mid_value=0.5, mid_color="FEF3C7",
                       end_type="num", end_value=1, end_color="A7F3D0"),
    )
    sheet.conditional_formatting.add(
        span("Drawdown"),
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="FCA5A5", showValue=True),
    )
    sheet.conditional_formatting.add(
        span("Risk % bal."),
        CellIsRule(operator="greaterThan", formula=["2"],
                   fill=_fill(AMBER_SOFT), font=Font(color="92400E", bold=True)),
    )


# -- dashboard -------------------------------------------------------------
def _kpi(sheet, row: int, column: int, label: str, value, note: str = "",
         number_format: str | None = None, tone: str | None = None) -> None:
    """A three-row KPI card: label, big value, footnote."""
    end = column + 2
    for offset in range(3):
        sheet.merge_cells(start_row=row + offset, start_column=column,
                          end_row=row + offset, end_column=end)
    _write(sheet, row, column, label.upper(), "ps_kpi_label")
    cell = _write(sheet, row + 1, column, value if value is not None else "n/a",
                  "ps_kpi_value", number_format)
    if tone:
        cell.font = Font(name="Segoe UI Semibold", size=16, color=tone)
    _write(sheet, row + 2, column, note, "ps_kpi_note")
    for offset in range(3):
        for col in range(column, end + 1):
            sheet.cell(row=row + offset, column=col).fill = _fill(CARD)
    edge = Side(style="thin", color=LINE)
    for col in range(column, end + 1):
        sheet.cell(row=row, column=col).border = Border(top=edge)
        sheet.cell(row=row + 2, column=col).border = Border(bottom=edge)
    sheet.cell(row=row + 1, column=column).border = Border(left=edge)
    sheet.cell(row=row + 1, column=end).border = Border(right=edge)
    sheet.row_dimensions[row].height = 15
    sheet.row_dimensions[row + 1].height = 26
    sheet.row_dimensions[row + 2].height = 13


def _tone(value, invert: bool = False) -> str | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number == 0:
        return SLATE
    good = number < 0 if invert else number > 0
    return TEAL if good else RED


def _fmt_duration(seconds) -> str:
    """Seconds / minutes / hours+minutes, exactly as the brief specified."""
    total = _duration_seconds(seconds)
    if total is None:
        return "-"
    threshold = int(settings.get("analysis.seconds_threshold", 61))
    if total < threshold:
        return f"{total:.0f}s"
    minutes = total / 60.0
    if minutes < 60:
        return f"{minutes:.0f}m"
    hours = int(total // 3600)
    remainder = int(round((total - hours * 3600) / 60))
    if remainder == 60:
        hours, remainder = hours + 1, 0
    return f"{hours}h {remainder:02d}m"


def build_dashboard(book: Workbook, positions: list[dict], summary: dict,
                    account: dict, monte: dict, first_row: int,
                    last_row: int) -> None:
    sheet = book.create_sheet("Dashboard", 0)
    login = account.get("login", "")
    server = account.get("server") or ""
    currency = account.get("currency") or ""
    _banner(
        sheet,
        f"{paths.APP_TITLE} — Mentorship Report",
        f"Account {login} · {server} · {currency} · "
        f"{summary.get('trades', 0)} closed trades · generated "
        f"{dt.datetime.now():%d %b %Y %H:%M}",
        width=16,
    )
    for column in range(1, 17):
        sheet.column_dimensions[get_column_letter(column)].width = 11.5
    # The chart helper tables live out here and are hidden: their numbers are
    # already on the face of the dashboard, and a stray two-column table beside
    # a pie is the sort of thing that makes a report look unfinished.
    for column in (HELPER_COL, HELPER_COL + 1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = 14
        sheet.column_dimensions[letter].hidden = True

    if not summary.get("sufficient"):
        _write(
            sheet, 3, 1,
            f"Sample is {summary.get('trades', 0)} trades. Distribution "
            f"statistics (SQN, Monte Carlo) need {stats.MIN_SAMPLE}+ to mean "
            "anything and are labelled where shown.",
            "ps_note",
        )
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=16)

    net = summary.get("net_profit")
    _write(sheet, 5, 1, "Performance", "ps_section")
    cards = [
        ("Net profit", net, f"{currency} · gross {summary.get('gross_profit')}",
         MONEY_SIGN, _tone(net)),
        ("Return", summary.get("return_pct"),
         f"on {summary.get('start_balance')} start", PCT, _tone(summary.get("return_pct"))),
        ("Win rate", summary.get("win_rate"),
         f"{summary.get('wins')}W / {summary.get('losses')}L", PCT, None),
        ("Profit factor", summary.get("profit_factor"),
         "gross win ÷ gross loss", RATIO,
         _tone((summary.get("profit_factor") or 1) - 1)),
        ("Expectancy", summary.get("expectancy"),
         f"per trade · {summary.get('expectancy_r')}R", MONEY_SIGN,
         _tone(summary.get("expectancy"))),
    ]
    for index, (label, value, note, number_format, tone) in enumerate(cards):
        _kpi(sheet, 6, 1 + index * 3, label, value, note, number_format, tone)

    _write(sheet, 10, 1, "Risk", "ps_section")
    dd = summary.get("max_drawdown")
    cards = [
        ("Max drawdown", -dd if dd else dd,
         f"{summary.get('max_drawdown_pct')}% of peak", MONEY, RED if dd else None),
        ("Recovery factor", summary.get("recovery_factor"),
         "net ÷ max drawdown", RATIO, None),
        ("Payoff ratio", summary.get("payoff_ratio"),
         "avg win ÷ avg loss", RATIO, None),
        ("SQN", summary.get("sqn"),
         summary.get("sqn_grade") or f"needs {stats.MIN_SAMPLE}+ trades", RATIO, None),
        ("Worst streak", summary.get("max_loss_streak"),
         f"best {summary.get('max_win_streak')} · "
         f"{summary.get('worst_run_money')} deepest run", INT, None),
    ]
    for index, (label, value, note, number_format, tone) in enumerate(cards):
        _kpi(sheet, 11, 1 + index * 3, label, value, note, number_format, tone)

    _write(sheet, 15, 1, "Trade management", "ps_section")
    cards = [
        ("Avg MFE", summary.get("avg_mfe"), "money offered", MONEY, TEAL),
        ("Avg MAE", summary.get("avg_mae"), "heat taken", MONEY, RED),
        ("Winner capture", summary.get("capture_of_winners"),
         f"median {summary.get('median_winner_capture')}", RATIO, None),
        ("Median heat", summary.get("median_heat_pct"),
         f"p90 {summary.get('p90_heat_pct')}%", PCT, None),
        ("Gave back", summary.get("reversed_winners"),
         f"{summary.get('reversed_winners_pct')}% of losers reached "
         f"{summary.get('reversed_winners_threshold_r')}R", INT, AMBER),
    ]
    for index, (label, value, note, number_format, tone) in enumerate(cards):
        _kpi(sheet, 16, 1 + index * 3, label, value, note, number_format, tone)

    _write(sheet, 20, 1, "Duration", "ps_section")
    cards = [
        ("Avg duration", _fmt_duration(summary.get("avg_duration_s")),
         f"{summary.get('avg_duration_s')} seconds", None, None),
        ("Median duration", _fmt_duration(summary.get("median_duration_s")),
         "half the trades are shorter", None, None),
        ("Avg risk", summary.get("avg_risk_money"),
         f"{summary.get('avg_risk_pct')}% of balance", MONEY, None),
        ("Risk consistency", summary.get("risk_consistency"),
         "stdev ÷ mean · lower is steadier", RATIO, None),
        ("Planned RR", summary.get("avg_planned_rr"),
         f"realised {summary.get('expectancy_r')}R", RATIO, None),
    ]
    for index, (label, value, note, number_format, tone) in enumerate(cards):
        _kpi(sheet, 21, 1 + index * 3, label, value, note, number_format, tone)

    _write(sheet, 25, 1, f"Equity curve · balance in {currency}", "ps_section")
    _write(sheet, 25, 11, "Outcome split", "ps_section")
    _chart_equity(sheet, first_row, last_row)
    _chart_outcomes(sheet, summary)
    _write(sheet, 44, 1, "R-multiple distribution · trades per band", "ps_section")
    _write(sheet, 44, 11, "MFE offered (up) vs MAE heat taken (right)", "ps_section")
    _chart_r_distribution(sheet, positions)
    _chart_excursion(sheet, first_row, last_row)
    _monte_carlo_block(sheet, 63, monte, currency)
    _write(
        sheet, 72, 1,
        "Every figure on this sheet is computed from the Trades sheet, which is "
        "the only source of data in this workbook. Change a filter there and the "
        "pivot sheets follow; the cards above are point-in-time values from the "
        "run that generated the file.",
        "ps_note",
    )
    sheet.merge_cells(start_row=72, start_column=1, end_row=73, end_column=16)
    # Break above the lower chart row so neither chart prints with its plot area
    # sliced in half across two sheets of paper.
    sheet.row_breaks.append(Break(id=43))


def _style_chart(chart, title: str | None, height: float = 8.5,
                 width: float = 21.0) -> None:
    """Common chart chrome. `title=None` leaves the chart untitled.

    On the dashboard the section heading above each chart already names it, and a
    chart title there just repeats itself and crowds the plot.
    """
    if title:
        chart.title = title
    chart.height = height
    chart.width = width
    chart.style = None
    # Excel defaults to plotting visible cells only, which would empty any chart
    # fed from the hidden helper columns. The ranges are chosen deliberately, so
    # the chart plots them whether or not their column is on screen.
    chart.visible_cells_only = False
    # A pie has no axes at all, so this is asked rather than assumed.
    if getattr(chart, "y_axis", None) is not None:
        chart.y_axis.majorGridlines.spPr = None
        chart.y_axis.delete = False
    if getattr(chart, "x_axis", None) is not None:
        chart.x_axis.delete = False


def _text_categories(chart, sheet_title: str, column: int,
                     first_row: int, last_row: int) -> None:
    """Point a chart's categories at a range of text labels.

    openpyxl's `set_categories` always writes a numeric reference. Excel then
    has to guess what to do with a column of words; pies and bar charts end up
    labelled 1, 2, 3. Writing the reference as a string source is what makes the
    labels come through as written.
    """
    letter = get_column_letter(column)
    span = absolute_coordinate(f"{letter}{first_row}:{letter}{last_row}")
    ref = f"{quote_sheetname(sheet_title)}!{span}"
    for series in chart.series:
        series.cat = AxDataSource(strRef=StrRef(f=ref))


def _chart_equity(sheet, first_row: int, last_row: int) -> None:
    """Running balance straight off the Trades table - always in step with it."""
    chart = LineChart()
    _style_chart(chart, None)
    trades = sheet.parent[SHEET_TRADES]
    data = Reference(trades, min_col=_col_index("Balance after"),
                     min_row=first_row - 1, max_row=last_row)
    labels = Reference(trades, min_col=_col_index("#"),
                       min_row=first_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    series = chart.series[0]
    series.graphicalProperties.line = LineProperties(solidFill=BLUE, w=22000)
    series.smooth = False
    # No axis title: 73 tick labels already fill the strip under the plot and a
    # title there lands on top of them.
    # One series needs no legend, and Excel fills a legend it was given with an
    # entry per category rather than leaving it empty.
    chart.legend = None
    sheet.add_chart(chart, "A26")


def _chart_outcomes(sheet, summary: dict) -> None:
    """Win / loss / breakeven counts.

    Chart series carry literal values rather than formulas: a pie fed by
    COUNTIF has no cached result until Excel recalculates, and a workbook that
    opens with an empty chart looks broken even when it is not.
    """
    anchor = 25
    _write(sheet, anchor, HELPER_COL, "Outcome", "ps_header")
    _write(sheet, anchor, HELPER_COL + 1, "Trades", "ps_header")
    rows = [
        ("Wins", summary.get("wins") or 0),
        ("Losses", summary.get("losses") or 0),
        ("Breakeven", summary.get("scratches") or 0),
    ]
    for offset, (label, value) in enumerate(rows, start=1):
        _write(sheet, anchor + offset, HELPER_COL, label, "ps_cell_left")
        _write(sheet, anchor + offset, HELPER_COL + 1, value, "ps_cell", INT)
    chart = PieChart()
    _style_chart(chart, None, height=8.5, width=11.0)
    chart.add_data(
        Reference(sheet, min_col=HELPER_COL + 1, min_row=anchor,
                  max_row=anchor + len(rows)),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=HELPER_COL, min_row=anchor + 1,
                  max_row=anchor + len(rows))
    )
    _text_categories(chart, sheet.title, HELPER_COL, anchor + 1, anchor + len(rows))
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    # Value and category are already carried by the legend and the percentage,
    # so switching them off is what keeps three slices readable. The legend key
    # would otherwise print a small swatch in front of every percentage.
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.showBubbleSize = False
    # Win blue, loss red, breakeven teal - the same colours the rest of the
    # workbook uses for those outcomes, so the pie needs no learning.
    for index, tone in enumerate((BLUE, RED, TEAL)):
        point = DataPoint(idx=index)
        point.graphicalProperties.solidFill = tone
        point.graphicalProperties.line.solidFill = WHITE
        chart.series[0].data_points.append(point)
    sheet.add_chart(chart, "K26")


def _col_index(header: str) -> int:
    for index, (title, *_rest) in enumerate(COLUMNS, start=1):
        if title == header:
            return index
    raise KeyError(header)


def _bucket_counts(positions: list[dict], key: str) -> list[tuple[str, int, float]]:
    """(label, trades, net) per bucket, ordered by the bucket's own prefix.

    model.r_bucket and model.duration_bucket prefix their labels A..J precisely
    so a plain sort puts them in trading order rather than alphabetical order.
    """
    tally: dict[str, list[float]] = {}
    for position in positions:
        label = (position.get("metrics") or {}).get(key)
        if not label:
            continue
        tally.setdefault(str(label), []).append(
            float(position.get("net_profit") or 0.0)
        )
    return [
        (label, len(nets), round(sum(nets), 2))
        for label, nets in sorted(tally.items())
    ]


def _chart_r_distribution(sheet, positions: list[dict]) -> None:
    """How the R outcomes are actually distributed, not just their average."""
    anchor = 44
    _write(sheet, anchor, HELPER_COL, "R band", "ps_header")
    _write(sheet, anchor, HELPER_COL + 1, "Trades", "ps_header")
    rows = _bucket_counts(positions, "r_bucket")
    for offset, (label, count, _net) in enumerate(rows, start=1):
        _write(sheet, anchor + offset, HELPER_COL,
               label[3:] if label[1:3] == ". " else label, "ps_cell_left")
        _write(sheet, anchor + offset, HELPER_COL + 1, count, "ps_cell", INT)
    if not rows:
        return
    chart = BarChart()
    chart.type = "col"
    _style_chart(chart, None)
    chart.add_data(
        Reference(sheet, min_col=HELPER_COL + 1, min_row=anchor,
                  max_row=anchor + len(rows)),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=HELPER_COL, min_row=anchor + 1,
                  max_row=anchor + len(rows))
    )
    _text_categories(chart, sheet.title, HELPER_COL, anchor + 1, anchor + len(rows))
    chart.series[0].graphicalProperties.solidFill = BLUE
    chart.gapWidth = 40
    chart.legend = None
    sheet.add_chart(chart, "A45")


def _chart_excursion(sheet, first_row: int, last_row: int) -> None:
    """MAE on x, MFE on y. Points hugging the x axis are trades that never went
    the trader's way; points far up and right are trades that paid but hurt."""
    trades = sheet.parent[SHEET_TRADES]
    chart = ScatterChart()
    chart.scatterStyle = "marker"
    _style_chart(chart, None, height=8.5, width=11.0)
    xvalues = Reference(trades, min_col=_col_index("MAE"),
                        min_row=first_row, max_row=last_row)
    yvalues = Reference(trades, min_col=_col_index("MFE"),
                        min_row=first_row - 1, max_row=last_row)
    series = Series(yvalues, xvalues, title_from_data=True)
    series.marker = Marker(symbol="circle", size=6)
    series.marker.graphicalProperties.solidFill = BLUE
    series.marker.graphicalProperties.line.solidFill = WHITE
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    # One series of 73 trades is one population, not 73 categories, so it gets
    # one colour.
    chart.varyColors = False
    # Axis titles are in the section heading instead: with the labels pinned low
    # an axis title lands on top of them, and there is no room to move it.
    # MAE is negative money, so the axes cross to the right of every point and
    # Excel would otherwise stack the tick labels on top of the data. "low" pins
    # them to the edge of the plot where they belong.
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"
    chart.legend = None
    sheet.add_chart(chart, "K45")


def _monte_carlo_block(sheet, row: int, monte: dict, currency: str) -> None:
    """Resampled forward projection, or a plain statement of why there isn't one."""
    _write(sheet, row, 1, "Forward risk (Monte Carlo)", "ps_section")
    if not monte.get("available"):
        _write(sheet, row + 1, 1, monte.get("reason", "Not enough history."),
               "ps_note")
        sheet.merge_cells(start_row=row + 1, start_column=1,
                          end_row=row + 1, end_column=16)
        return
    _write(
        sheet, row + 1, 1,
        f"{monte['runs']:,} runs of {monte['horizon']} trades, drawn with "
        f"replacement from this account's own {monte['sample']} results - so the "
        "tails in the real record stay in the projection.",
        "ps_note",
    )
    sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=16)
    cards = [
        ("Median outcome", monte.get("median_final"),
         f"balance after {monte['horizon']} trades", MONEY, None),
        ("Bad case (5%)", monte.get("p05_final"),
         "1 run in 20 ends here or worse", MONEY, RED),
        ("Good case (95%)", monte.get("p95_final"),
         "1 run in 20 ends here or better", MONEY, TEAL),
        ("Drawdown to expect", monte.get("median_drawdown"),
         f"p95 {monte.get('p95_drawdown')} · worst "
         f"{monte.get('worst_drawdown')}", MONEY, None),
        ("Chance of profit", monte.get("profitable_pct"),
         f"risk of halving the account {monte.get('risk_of_50pct_loss')}%",
         PCT, None),
    ]
    for index, (label, value, note, number_format, tone) in enumerate(cards):
        _kpi(sheet, row + 3, 1 + index * 3, label, value, note, number_format, tone)


# -- pivot sheets ----------------------------------------------------------
# Column layout shared by every pivot grid, so the sheets read the same way.
PIVOT_COLUMNS = [
    ("Trades", 9), ("Wins", 8), ("Win %", 9), ("Net", 12),
    ("Expectancy", 11), ("Avg R", 9), ("Profit factor", 12),
    ("Avg MFE", 10), ("Avg MAE", 10), ("Avg dur (min)", 13),
    ("Share of net", 12),
]

FILTER_ROW = 4


def _slicers(sheet, symbols: list[str], first: int, last: int,
             row: int = FILTER_ROW) -> str:
    """Two dropdown filters, and the SUMIFS criteria fragment they drive.

    "(All)" becomes the wildcard "*", which matches any text - Symbol and
    Outcome are always populated, so nothing is silently dropped.
    """
    _write(sheet, row, 1, "Filters", "ps_section")
    _write(sheet, row, 3, "Symbol", "ps_label")
    _write(sheet, row, 6, "Outcome", "ps_label")
    symbol_cell = _write(sheet, row, 4, "(All)", "ps_cell",
                         fill=AMBER_SOFT, border=_thin(AMBER))
    outcome_cell = _write(sheet, row, 7, "(All)", "ps_cell",
                          fill=AMBER_SOFT, border=_thin(AMBER))
    symbol_list = ",".join(["(All)"] + [str(s) for s in symbols])
    if len(symbol_list) <= 250:
        rule = DataValidation(type="list", formula1=f'"{symbol_list}"',
                              allow_blank=False, showDropDown=False)
        sheet.add_data_validation(rule)
        rule.add(symbol_cell)
    outcomes = DataValidation(type="list",
                              formula1='"(All),Win,Loss,Breakeven"',
                              allow_blank=False, showDropDown=False)
    sheet.add_data_validation(outcomes)
    outcomes.add(outcome_cell)
    _write(sheet, row, 9,
           "Change either dropdown - every grid below recalculates.",
           "ps_note_flow")
    return (
        f",{rng('Symbol', first, last)},IF($D${row}=\"(All)\",\"*\",$D${row})"
        f",{rng('Outcome', first, last)},IF($G${row}=\"(All)\",\"*\",$G${row})"
    )


PIVOT_FORMATS = [INT, INT, PCT, MONEY_SIGN, MONEY_SIGN, '+0.00;[Red]-0.00;0.00',
                 RATIO, MONEY, MONEY, "0.0", PCT]


def _pivot_grid(sheet, row: int, title: str, dimension: str, values: list[str],
                first: int, last: int, criteria: str, note: str = "") -> int:
    """One aggregation grid. Returns the next free row.

    Every cell is a formula over the Trades table, so the grid is auditable and
    responds to the slicers above it - which is the part a cached pivot table
    could not do without a manual refresh.
    """
    net = rng("Net", first, last)
    dim = rng(dimension, first, last)
    outcome = rng("Outcome", first, last)

    _write(sheet, row, 1, title, "ps_section")
    if note:
        _write(sheet, row, 5, note, "ps_note_flow")
    header = row + 1
    _write(sheet, header, 1, dimension, "ps_header")
    for index, (label, width) in enumerate(PIVOT_COLUMNS, start=2):
        _write(sheet, header, index, label, "ps_header")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[header].height = 28

    def formulas(key: str, dim_criteria: str) -> list[str]:
        r = key
        return [
            f"=COUNTIFS({dim_criteria[1:]}{criteria})" if dim_criteria
            else f"=COUNTIFS({criteria[1:]})",
            f"=COUNTIFS({outcome},\"Win\"{dim_criteria}{criteria})",
            f'=IFERROR(C{r}/B{r}*100,"")',
            f"=SUMIFS({net}{dim_criteria}{criteria})",
            f'=IFERROR(E{r}/B{r},"")',
            f'=IFERROR(AVERAGEIFS({rng("R multiple", first, last)}'
            f'{dim_criteria}{criteria}),"")',
            f'=IFERROR(SUMIFS({net}{dim_criteria},{net},">0"{criteria})'
            f'/-SUMIFS({net}{dim_criteria},{net},"<0"{criteria}),"")',
            f'=IFERROR(AVERAGEIFS({rng("MFE", first, last)}'
            f'{dim_criteria}{criteria}),"")',
            f'=IFERROR(AVERAGEIFS({rng("MAE", first, last)}'
            f'{dim_criteria}{criteria}),"")',
            f'=IFERROR(AVERAGEIFS({rng("Duration (min)", first, last)}'
            f'{dim_criteria}{criteria}),"")',
            f'=IFERROR(E{r}/SUMIFS({net}{criteria})*100,"")',
        ]

    for offset, value in enumerate(values):
        line = header + 1 + offset
        _write(sheet, line, 1, value, "ps_cell_left")
        dim_criteria = f",{dim},$A${line}"
        for index, formula in enumerate(formulas(str(line), dim_criteria), start=2):
            _write(sheet, line, index, formula, "ps_cell",
                   PIVOT_FORMATS[index - 2])
        sheet.row_dimensions[line].height = 16

    total = header + 1 + len(values)
    _write(sheet, total, 1, "All trades", "ps_cell_left", fill=CARD,
           font=Font(name="Segoe UI Semibold", size=9, color=INK))
    for index, formula in enumerate(formulas(str(total), ""), start=2):
        cell = _write(sheet, total, index, formula, "ps_cell",
                      PIVOT_FORMATS[index - 2])
        cell.font = Font(name="Segoe UI Semibold", size=9, color=INK)
        cell.fill = _fill(CARD)

    if values:
        body = f"E{header + 1}:E{total - 1}"
        sheet.conditional_formatting.add(
            body,
            CellIsRule(operator="greaterThan", formula=["0"],
                       font=Font(color="065F46", bold=True)),
        )
        sheet.conditional_formatting.add(
            body,
            CellIsRule(operator="lessThan", formula=["0"],
                       font=Font(color="991B1B", bold=True)),
        )
        sheet.conditional_formatting.add(
            f"B{header + 1}:B{total - 1}",
            DataBarRule(start_type="num", start_value=0, end_type="max",
                        color=BLUE_SOFT, showValue=True),
        )
    return total + 3


def _values_of(positions: list[dict], header: str) -> list:
    """Distinct values of a Trades column, in the order the grid should show."""
    _title, key, source, _fmt, _width = next(
        entry for entry in COLUMNS if entry[0] == header
    )
    found = set()
    for position in positions:
        value = _cell_value(position, key, source, {}, header)
        if value is None or value == "":
            continue
        found.add(value)
    numeric = all(isinstance(v, (int, float)) and not isinstance(v, bool)
                  for v in found)
    return sorted(found) if numeric else sorted(str(v) for v in found)


def _pivot_sheet(book: Workbook, name: str, title: str, subtitle: str,
                 positions: list[dict], dimensions: list[tuple[str, str]],
                 first: int, last: int):
    sheet = book.create_sheet(name)
    _banner(sheet, title, subtitle, width=12)
    sheet.column_dimensions["A"].width = 22
    criteria = _slicers(sheet, _values_of(positions, "Symbol"), first, last)
    row = FILTER_ROW + 2
    for header, note in dimensions:
        row = _pivot_grid(sheet, row, f"By {header.lower()}", header,
                          _values_of(positions, header), first, last,
                          criteria, note)
    sheet.freeze_panes = f"A{FILTER_ROW + 1}"
    return sheet


def build_pivots(book: Workbook, positions: list[dict], first: int,
                 last: int) -> None:
    _pivot_sheet(
        book, "Pivots", "What is making and losing the money",
        "Live formula grids over the Trades table. Both dropdowns apply to "
        "every grid on the sheet.",
        positions,
        [
            ("Symbol", "Concentration check: one instrument carrying the "
                       "account is a risk, not a strength."),
            ("Side", "A large gap between buy and sell says the edge is "
                     "directional, not systematic."),
            ("Exit reason", "How trades actually end, versus how they were "
                            "planned to end."),
            ("SL moved", "Whether moving a stop after entry helped or cost."),
        ],
        first, last,
    )
    _pivot_sheet(
        book, "Timing", "When the edge shows up",
        "Same grids, cut by clock and calendar. Broker server time throughout.",
        positions,
        [
            ("Session", "Sessions come from settings, in broker hours."),
            ("Weekday", ""),
            ("Hour", "Hour of entry. Thin rows are noise - read the trade "
                     "count before the net."),
            ("Month", ""),
        ],
        first, last,
    )
    _pivot_sheet(
        book, "Shape of results", "Distribution of outcomes",
        "Where the results sit, not just their average.",
        positions,
        [
            ("R band", "The distribution behind the expectancy figure."),
            ("Duration band", "Time in the market against money made."),
            ("MFE first", "Did price go the trader's way before it went "
                          "against? The 'No' row is trades that were wrong "
                          "from the first tick."),
        ],
        first, last,
    )


# -- excursions ------------------------------------------------------------
def _metric_row(sheet, row: int, label: str, value, note: str,
                number_format: str | None = None, tone: str | None = None) -> None:
    _write(sheet, row, 1, label, "ps_cell_left")
    cell = _write(sheet, row, 4, value if value is not None else "n/a",
                  "ps_cell", number_format)
    if tone:
        cell.font = Font(name="Segoe UI Semibold", size=9, color=tone)
    cell.border = _thin()
    _write(sheet, row, 6, note, "ps_note_flow")
    sheet.row_dimensions[row].height = 16


def build_excursions(book: Workbook, positions: list[dict], summary: dict,
                     first: int, last: int) -> None:
    """MAE and MFE, and the questions only they can answer."""
    sheet = book.create_sheet("Excursions")
    _banner(
        sheet, "Excursions — MAE, MFE and what happened in between",
        "Maximum adverse and favourable excursion measured tick by tick where "
        "the history allowed it, from M1 bars otherwise. The source is recorded "
        "per trade on the Trades sheet.",
        width=12,
    )
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["D"].width = 13
    for column in "BCE":
        sheet.column_dimensions[column].width = 4
    sheet.column_dimensions["F"].width = 70

    _write(sheet, 4, 1, "Headline", "ps_section")
    rows = [
        ("Average MFE", summary.get("avg_mfe"),
         "How much the average trade was up at its best point.", MONEY, TEAL),
        ("Average MAE", summary.get("avg_mae"),
         "How far the average trade went against the entry before resolving.",
         MONEY, RED),
        ("Average MAE on winners", summary.get("avg_winner_mae"),
         "Heat a winning trade takes. If this is far inside the average risk, "
         "the stops are wider than the trades need.", MONEY, None),
        ("Average MFE on losers", summary.get("avg_loser_mfe"),
         "Money that was on the table in trades that finished red.", MONEY, None),
        ("Median heat", summary.get("median_heat_pct"),
         "MAE as a share of the risk taken. Above 100% means the trade went "
         "further against the entry than the stop it was risking.", PCT, None),
        ("90th percentile heat", summary.get("p90_heat_pct"),
         "One trade in ten runs at least this hot.", PCT, None),
        ("Capture on winners", summary.get("capture_of_winners"),
         "Money kept divided by money offered, across all winners. 1.00 would "
         "be exiting every winner at its exact high.", RATIO, None),
        ("Median winner capture", summary.get("median_winner_capture"),
         "The typical winner, unweighted by size.", RATIO, None),
        ("Total giveback", summary.get("total_giveback"),
         "Sum of the gap between each trade's best point and where it closed.",
         MONEY, AMBER),
        ("Losers that were up "
         f"{summary.get('reversed_winners_threshold_r')}R or more",
         summary.get("reversed_winners"),
         f"{summary.get('reversed_winners_pct')}% of all losing trades. The "
         "threshold exists on purpose: measured on ticks almost every trade "
         "shows a fractionally positive MFE, and counting those would turn "
         "noise into a finding.", INT, AMBER),
        ("Trades with excursion data", summary.get("excursion_sample"),
         f"out of {summary.get('trades')} closed trades.", INT, None),
    ]
    for offset, (label, value, note, number_format, tone) in enumerate(rows):
        _metric_row(sheet, 5 + offset, label, value, note, number_format, tone)

    row = 5 + len(rows) + 2
    _write(sheet, row, 1, "Efficiency (Sweeney)", "ps_section")
    _write(sheet, row, 6,
           "Share of the trade's own high-to-low range that the entry, the exit "
           "and the trade as a whole captured. 1.00 is perfect, 0.50 is a coin "
           "flip against the range that was available.", "ps_note_flow")
    row += 1
    efficiency = [
        ("Entry efficiency", "Entry eff."),
        ("Exit efficiency", "Exit eff."),
        ("Total efficiency", "Total eff."),
    ]
    for offset, (label, header) in enumerate(efficiency):
        line = row + offset
        _write(sheet, line, 1, label, "ps_cell_left")
        _write(sheet, line, 4,
               f'=IFERROR(AVERAGE({rng(header, first, last)}),"")',
               "ps_cell", RATIO).border = _thin()
        _write(sheet, line, 6,
               f"Average across all closed trades, from column "
               f"{col_of(header)} of the Trades sheet.", "ps_note_flow")

    row += len(efficiency) + 2
    criteria = _slicers(sheet, _values_of(positions, "Symbol"), first, last, row)
    _pivot_grid(
        sheet, row + 2, "Excursion by exit reason", "Exit reason",
        _values_of(positions, "Exit reason"), first, last, criteria,
        "Compare Avg MFE against Avg MAE row by row: a reason whose MFE dwarfs "
        "its net is an exit problem, not an entry problem.",
    )


# -- risk and sequence -----------------------------------------------------
def build_risk(book: Workbook, positions: list[dict], summary: dict,
               first: int, last: int) -> None:
    """Position sizing, drawdown and the sequence the results arrived in."""
    sheet = book.create_sheet("Risk")
    _banner(
        sheet, "Risk and sequence",
        "Sizing consistency, drawdown depth, and what streaks did to the "
        "balance. Sequence matters: the same trades in a different order "
        "produce a different drawdown.",
        width=12,
    )
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["D"].width = 13
    for column in "BCE":
        sheet.column_dimensions[column].width = 4
    sheet.column_dimensions["F"].width = 70

    _write(sheet, 4, 1, "Sizing", "ps_section")
    rows = [
        ("Average risk", summary.get("avg_risk_money"),
         "Money at stake per trade, from entry to the initial stop.",
         MONEY, None),
        ("Median risk", summary.get("median_risk_money"),
         "Half the trades risked less than this.", MONEY, None),
        ("Risk consistency", summary.get("risk_consistency"),
         "Standard deviation divided by the mean. Under 0.30 is disciplined "
         "sizing; above 0.60 means the size is being chosen by feel, and one "
         "oversized loss can undo a good month.", RATIO, None),
        ("Average risk % of balance", summary.get("avg_risk_pct"),
         "Against the balance standing before each trade.", PCT, None),
        ("Largest risk % of balance", summary.get("max_risk_pct"),
         "The single most exposed trade in the record.", PCT, AMBER),
        ("Average planned RR", summary.get("avg_planned_rr"),
         "Reward-to-risk as set at entry, where both a stop and a target "
         "existed.", RATIO, None),
        ("Realised R per trade", summary.get("expectancy_r"),
         "Compare with planned RR above: a wide gap is trades not being held "
         "to their own plan.", RATIO, _tone(summary.get("expectancy_r"))),
    ]
    for offset, (label, value, note, number_format, tone) in enumerate(rows):
        _metric_row(sheet, 5 + offset, label, value, note, number_format, tone)

    row = 5 + len(rows) + 2
    _write(sheet, row, 1, "Drawdown", "ps_section")
    dd = summary.get("max_drawdown")
    rows = [
        ("Maximum drawdown", -dd if dd else dd,
         f"{summary.get('max_drawdown_pct')}% below the equity peak.",
         MONEY, RED if dd else None),
        ("Recovery factor", summary.get("recovery_factor"),
         "Net profit divided by that drawdown. Under 1.0 means the account has "
         "not yet made back more than its worst decline.", RATIO, None),
        ("Longest drawdown", summary.get("longest_dd_trades"),
         "Trades spent below the previous equity peak - the number that "
         "actually tests patience.", INT, None),
        ("Worst losing run", summary.get("worst_run_money"),
         f"{summary.get('max_loss_streak')} consecutive losses.", MONEY, RED),
        ("Best winning run", summary.get("best_run_money"),
         f"{summary.get('max_win_streak')} consecutive wins.", MONEY, TEAL),
        ("Current streak", summary.get("current_streak"),
         "Positive is wins, negative is losses, as of the last closed trade.",
         '+0;-0;0', None),
        ("Standard deviation per trade", summary.get("stdev_trade"),
         "The spread the expectancy is drawn from. A large value next to a "
         "small expectancy is why the equity curve feels rough.", MONEY, None),
        ("Largest win", summary.get("largest_win"),
         "Check this against average win: if one trade carries the account, "
         "the record is thinner than the net suggests.", MONEY, TEAL),
        ("Largest loss", summary.get("largest_loss"),
         "Against average risk, this shows whether stops held.", MONEY, RED),
    ]
    for offset, (label, value, note, number_format, tone) in enumerate(rows):
        _metric_row(sheet, row + 1 + offset, label, value, note, number_format, tone)

    row = row + 1 + len(rows) + 2
    criteria = _slicers(sheet, _values_of(positions, "Symbol"), first, last, row)
    row = _pivot_grid(
        sheet, row + 2, "By stop management", "SL moved",
        _values_of(positions, "SL moved"), first, last, criteria,
        "Whether moving the stop after entry paid for itself. 'Tightened' with "
        "a worse net than 'Unchanged' means stops are being moved too early.",
    )
    row = _pivot_grid(
        sheet, row, "By position in the streak", "Streak",
        _values_of(positions, "Streak"), first, last, criteria,
        "Streak counts the trade itself: -1 is the first loss after a win, -2 "
        "the second in a row. If the net worsens as the streak deepens, the "
        "losing runs are being made worse by the trading, not just by the "
        "market.",
    )
    _pivot_grid(
        sheet, row, "By near miss on the stop", "Near miss",
        _values_of(positions, "Near miss"), first, last, criteria,
        "'Yes' means price came within 5 points of the stop and did not take "
        "it. A pile of these says the stops are placed exactly where the market "
        "turns - which is luck, and it runs out.",
    )


# -- findings --------------------------------------------------------------
# A finding is only written when the evidence clears both a sample floor and an
# effect-size gate. Without those, a detector run over 73 trades will always
# produce a full page of confident-sounding text, because some subset of any
# record looks bad. The gates are stated next to each finding so the trader can
# judge the claim rather than take it on faith.
FINDING_MIN_SAMPLE = 8


def _finding(title: str, verdict: str, evidence: str, action: str,
             severity: str) -> dict:
    return {"title": title, "verdict": verdict, "evidence": evidence,
            "action": action, "severity": severity}


def _findings(positions: list[dict], summary: dict) -> list[dict]:
    """Behavioural findings, each gated on sample size and effect size."""
    found: list[dict] = []
    trades = summary.get("trades") or 0
    metrics = [p.get("metrics") or {} for p in positions]

    capture = summary.get("capture_of_winners")
    if capture is not None and summary.get("wins", 0) >= FINDING_MIN_SAMPLE:
        if capture < 0.7:
            found.append(_finding(
                "Winners are being closed well before their high",
                f"Winners kept {capture:.0%} of what they were offered.",
                f"Across {summary['wins']} winning trades, total giveback was "
                f"{summary.get('total_giveback')}. Gate: fewer than 70% "
                f"captured, at least {FINDING_MIN_SAMPLE} winners.",
                "Trail the stop behind structure instead of closing at a fixed "
                "distance, and compare the result over the next 20 trades.",
                "high" if capture < 0.5 else "medium",
            ))
        elif capture >= 0.9:
            found.append(_finding(
                "Exits on winners are efficient",
                f"Winners kept {capture:.0%} of what they were offered.",
                f"Median winner capture {summary.get('median_winner_capture')}. "
                "This is a strength worth protecting, not a problem.",
                "Leave the exit rule alone. If profit needs to grow, it has to "
                "come from entries or size, not from holding longer.",
                "good",
            ))

    heat = summary.get("median_heat_pct")
    sample = summary.get("excursion_sample") or 0
    if heat is not None and sample >= FINDING_MIN_SAMPLE and heat >= 90:
        found.append(_finding(
            "Trades routinely run most of the way to the stop",
            f"Median heat {heat}% of the risk taken; p90 "
            f"{summary.get('p90_heat_pct')}%.",
            f"Measured on {sample} trades with excursion data. Gate: median "
            "heat at or above 90%.",
            "Either the entries are early or the stops are tight. Check "
            "average MAE on winners on the Excursions sheet: if winners also "
            "run hot, the entries are the problem.",
            "high" if heat >= 100 else "medium",
        ))

    losers = summary.get("losses") or 0
    reversed_winners = summary.get("reversed_winners") or 0
    if losers >= FINDING_MIN_SAMPLE and reversed_winners:
        share = reversed_winners / losers
        if share >= 0.35:
            found.append(_finding(
                "Half-won trades are being allowed to become losses",
                f"{reversed_winners} of {losers} losing trades "
                f"({share:.0%}) had reached "
                f"{summary.get('reversed_winners_threshold_r')}R in profit.",
                f"Gate: at least 35% of losers, minimum "
                f"{FINDING_MIN_SAMPLE} losers, and each one had to reach "
                f"{summary.get('reversed_winners_threshold_r')}R - a lower bar "
                "would count tick noise as a giveback.",
                "A stop to breakeven once a trade reaches 1R would convert "
                "these to scratches. Test it on the Trades sheet before "
                "adopting it: it also cuts some eventual winners.",
                "high",
            ))

    consistency = summary.get("risk_consistency")
    if consistency is not None and trades >= FINDING_MIN_SAMPLE * 2:
        if consistency > 0.6:
            found.append(_finding(
                "Position size is inconsistent",
                f"Risk varies with a spread {consistency:.2f} times its own "
                "average.",
                f"Average risk {summary.get('avg_risk_money')}, largest single "
                f"exposure {summary.get('max_risk_pct')}% of balance. Gate: "
                "spread-to-mean above 0.60.",
                "Fix risk per trade as a percentage of balance and let the lot "
                "size follow from the stop distance.",
                "high" if consistency > 1.0 else "medium",
            ))
        elif consistency < 0.3:
            found.append(_finding(
                "Position sizing is consistent",
                f"Risk spread is only {consistency:.2f} of its average.",
                "Sizing is not the source of the volatility in this record.",
                "Keep it. It means the equity curve reflects the strategy "
                "rather than the size chosen on the day.",
                "good",
            ))

    payoff = summary.get("payoff_ratio")
    win_rate = summary.get("win_rate")
    if payoff and win_rate is not None and trades >= FINDING_MIN_SAMPLE * 2:
        breakeven = 100.0 / (1.0 + payoff)
        margin = win_rate - breakeven
        if 0 < margin < 6:
            found.append(_finding(
                "The edge is real but thin",
                f"Win rate {win_rate}% against a breakeven requirement of "
                f"{breakeven:.1f}% at a {payoff:.2f} payoff.",
                f"That leaves {margin:.1f} points of margin over "
                f"{trades} trades. Gate: positive but under 6 points.",
                "Widen the margin before adding size. Either the payoff or the "
                "win rate has to move; costs alone can erase this.",
                "medium",
            ))
        elif margin <= 0:
            found.append(_finding(
                "Win rate is below what the payoff needs",
                f"Win rate {win_rate}% against {breakeven:.1f}% required at a "
                f"{payoff:.2f} payoff.",
                f"Over {trades} trades. Any profit in the record came from a "
                "small number of outsized results, not from the base rate.",
                "Do not scale this. Raise the payoff by holding winners "
                "further, or tighten entry selection until the win rate clears "
                "the requirement.",
                "high",
            ))

    return found


def _group_findings(positions: list[dict], summary: dict) -> list[dict]:
    """Findings about a slice of the record - session, symbol, hour, duration.

    The gate here is deliberately harsh. With five sessions and 73 trades, the
    worst-performing session is worst by arithmetic necessity; that is not a
    finding. A slice only earns a line when it holds enough trades to mean
    something and moves the account's net by a material amount.
    """
    found: list[dict] = []
    net_total = summary.get("net_profit") or 0.0
    if not net_total:
        return found

    dimensions = [
        ("session", lambda p: (p.get("metrics") or {}).get("session"), "session"),
        ("symbol", lambda p: p.get("symbol"), "instrument"),
        ("weekday", lambda p: (p.get("metrics") or {}).get("weekday"), "weekday"),
        ("duration band",
         lambda p: (p.get("metrics") or {}).get("duration_bucket"), "holding time"),
    ]
    for label, key_fn, noun in dimensions:
        rows = stats.group_by(positions, key_fn, min_sample=FINDING_MIN_SAMPLE)
        if len(rows) < 2:
            continue
        worst = rows[-1]
        best = rows[0]
        # "Material" means the slice moved the account's net by at least a
        # quarter of the total. A 3% drag is inside the noise of 73 trades.
        if worst["net"] < 0 and abs(worst["net"]) >= abs(net_total) * 0.25:
            found.append(_finding(
                f"One {noun} is carrying most of the losses",
                f"{label.title()} \"{worst['key']}\" lost {worst['net']} across "
                f"{worst['trades']} trades, at a {worst['win_rate']}% win rate.",
                f"That is {abs(worst['net'] / net_total):.0%} of the account's "
                f"net result. Gate: at least {FINDING_MIN_SAMPLE} trades in the "
                "slice and a loss worth 25% or more of net.",
                f"Stop trading this {noun} for 20 trades and compare the "
                "account's net over that stretch. If it improves, the slice was "
                "the problem, not the sample.",
                "high",
            ))
        if best["net"] > 0 and best["net"] >= abs(net_total) * 0.4:
            found.append(_finding(
                f"One {noun} is producing most of the profit",
                f"{label.title()} \"{best['key']}\" made {best['net']} across "
                f"{best['trades']} trades, at a {best['win_rate']}% win rate.",
                f"That is {best['net'] / net_total:.0%} of net. Gate: at least "
                f"{FINDING_MIN_SAMPLE} trades and 40% or more of net. Read this "
                "as concentration, which cuts both ways.",
                f"Know that the account currently depends on this {noun}. Either "
                "deliberately specialise in it, or build a second edge before "
                "this one stops working.",
                "medium",
            ))
    return found


_SEVERITY = {
    "high": ("Act on this", RED, RED_SOFT),
    "medium": ("Worth watching", AMBER, AMBER_SOFT),
    "good": ("Strength", TEAL, TEAL_SOFT),
}


def build_findings(book: Workbook, positions: list[dict], summary: dict) -> None:
    sheet = book.create_sheet("Findings")
    _banner(
        sheet, "Findings",
        "Generated from this account's own record. Each finding states the "
        "threshold it had to clear, so a claim can be checked rather than "
        "believed.",
        width=12,
    )
    sheet.column_dimensions["A"].width = 4
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 46
    for column in "DEFGHIJKL":
        sheet.column_dimensions[column].width = 12

    found = _findings(positions, summary) + _group_findings(positions, summary)
    order = {"high": 0, "medium": 1, "good": 2}
    found.sort(key=lambda f: order.get(f["severity"], 3))

    row = 4
    if not found:
        _write(
            sheet, row, 2,
            "No finding cleared its threshold on this record.", "ps_section",
        )
        _write(
            sheet, row + 1, 2,
            f"That is a result, not an empty page. {summary.get('trades', 0)} "
            "trades were tested against every gate on this sheet - capture, "
            "heat, giveback, sizing consistency, edge margin, and per-session, "
            "per-symbol, per-weekday and per-duration concentration. Nothing "
            "was extreme enough, on enough trades, to be worth a "
            "recommendation. The pivot sheets still hold the underlying "
            "numbers if you want to look yourself.",
            "ps_note",
        )
        sheet.merge_cells(start_row=row + 1, start_column=2, end_row=row + 4,
                          end_column=12)
        return

    _write(sheet, row, 2,
           f"{len(found)} finding(s) cleared the evidence gates.", "ps_section")
    row += 2
    for index, finding in enumerate(found, start=1):
        badge, colour, soft = _SEVERITY.get(finding["severity"],
                                            ("Note", SLATE, CARD))
        _write(sheet, row, 1, index, "ps_cell", INT)
        _write(sheet, row, 2, badge, "ps_cell", fill=soft,
               font=Font(name="Segoe UI Semibold", size=9, color=colour),
               border=_thin(colour))
        _write(sheet, row, 3, finding["title"],
               font=Font(name="Segoe UI Semibold", size=11, color=INK),
               align=Alignment("left", "center"))
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        sheet.row_dimensions[row].height = 20

        for offset, (label, text) in enumerate(
            (("What the record shows", finding["verdict"]),
             ("Evidence and gate", finding["evidence"]),
             ("What to do", finding["action"])),
            start=1,
        ):
            line = row + offset
            _write(sheet, line, 2, label, "ps_label",
                   align=Alignment("right", "top"))
            _write(sheet, line, 3, text, "ps_note")
            sheet.merge_cells(start_row=line, start_column=3, end_row=line,
                              end_column=12)
            sheet.row_dimensions[line].height = 30
        for column in range(1, 13):
            sheet.cell(row=row + 4, column=column).border = Border(
                bottom=Side(style="thin", color=LINE)
            )
        row += 6


# -- charts index ----------------------------------------------------------
def build_charts_index(book: Workbook, positions: list[dict],
                       shots: dict[int, list[dict]],
                       folders: dict[int, str]) -> None:
    """One row per image, with a clickable link to the file on disk.

    Relative hyperlinks are used so the workbook keeps working if the Charts
    folder and the exported file are moved together.
    """
    sheet = book.create_sheet("Charts")
    _banner(
        sheet, "Chart index",
        "Every rendered shot, linked. Nine images per position: open and close "
        "at M15, M5 and M1, a single close at H1 and H4, and one entry-to-exit "
        "frame. Candles only - no indicators, no objects.",
        width=9,
    )
    widths = [8, 12, 10, 8, 8, 10, 12, 46, 30]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    headers = ["#", "Ticket", "Symbol", "Side", "TF", "Event", "Outcome",
               "Image", "Folder"]
    for index, title in enumerate(headers, start=1):
        _write(sheet, 4, index, title, "ps_header")
    sheet.row_dimensions[4].height = 26

    export_root = paths.exports_dir()
    row = 5
    for position in positions:
        entries = sorted(
            shots.get(position["id"], []),
            key=lambda s: (_SHOT_ORDER.get(s["timeframe"], 99), s["event"]),
        )
        metrics = position.get("metrics") or {}
        for shot in entries:
            target = paths.charts_dir() / shot["rel_path"]
            _write(sheet, row, 1, metrics.get("trade_no"), "ps_cell", INT)
            _write(sheet, row, 2, position.get("ticket"), "ps_cell", "0")
            _write(sheet, row, 3, position.get("symbol"), "ps_cell")
            _write(sheet, row, 4, position.get("side"), "ps_cell")
            _write(sheet, row, 5, shot["timeframe"], "ps_cell")
            _write(sheet, row, 6, shot["event"], "ps_cell")
            _write(sheet, row, 7, metrics.get("outcome"), "ps_cell")
            cell = _write(sheet, row, 8, Path(shot["rel_path"]).name,
                          "ps_cell_left")
            try:
                link = os.path.relpath(target, export_root).replace("\\", "/")
            except ValueError:  # different drive - fall back to absolute
                link = target.as_uri()
            cell.hyperlink = link
            cell.font = Font(name="Segoe UI", size=9, color=BLUE, underline="single")
            _write(sheet, row, 9, folders.get(position["id"], ""), "ps_cell_left")
            row += 1

    if row > 5:
        table = Table(displayName="ChartIndex", ref=f"A4:I{row - 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9", showRowStripes=True
        )
        sheet.add_table(table)
    sheet.freeze_panes = "A5"


_SHOT_ORDER = {"H4": 1, "H1": 2, "M15": 3, "M5": 4, "M1": 5}


# -- glossary --------------------------------------------------------------
GLOSSARY = [
    ("MAE", "Maximum adverse excursion",
     "The furthest a trade went against the entry before it closed, in account "
     "currency. It measures the pain a position caused, which the profit and "
     "loss figure hides entirely."),
    ("MFE", "Maximum favourable excursion",
     "The furthest a trade went in the trader's favour. The gap between MFE "
     "and the realised result is the money that was offered and not taken."),
    ("Heat %", "MAE as a share of risk",
     "MAE divided by the money risked at entry. 100% means the trade travelled "
     "the whole distance to the stop; above 100% means it went past where the "
     "stop was and came back, which usually means the stop was moved."),
    ("Capture", "Realised result ÷ MFE",
     "How much of the available move was kept. Only meaningful on a trade that "
     "had something to capture, so it is reported on winners."),
    ("Giveback", "MFE − realised result",
     "Money handed back after a trade was in profit."),
    ("R multiple", "Result ÷ risk",
     "The result expressed in units of the risk taken at entry. It makes a "
     "£20 win on £10 of risk and a £200 win on £100 of risk the same trade, "
     "which is the only way to compare across position sizes."),
    ("Planned RR", "Target distance ÷ stop distance",
     "The reward-to-risk the trade was set up to deliver, from the levels "
     "present at entry."),
    ("Expectancy", "Average result per trade",
     "What one trade is worth on average. Positive expectancy with a low win "
     "rate is normal and fine; negative expectancy with a high win rate is the "
     "classic profile of a strategy that will eventually give it all back."),
    ("Profit factor", "Gross profit ÷ gross loss",
     "Above 1.0 is profitable. Below about 1.2 the result is fragile to a "
     "change in costs or a run of bad luck."),
    ("Payoff ratio", "Average win ÷ average loss",
     "Together with the win rate it says whether the arithmetic works: the "
     "breakeven win rate is 1 ÷ (1 + payoff)."),
    ("SQN", "System Quality Number (Van Tharp)",
     "√n × expectancy ÷ standard deviation of results. It rewards consistency, "
     "not just profit. It is meaningless below about 20 trades and is withheld "
     "rather than shown when the sample is short."),
    ("Recovery factor", "Net profit ÷ maximum drawdown",
     "How much profit the account produced per unit of its worst decline."),
    ("Efficiency", "Sweeney entry, exit and total efficiency",
     "What share of the trade's own high-to-low range was captured by the "
     "entry, by the exit, and overall. 1.00 would be buying the low and selling "
     "the high."),
    ("Near miss", "Stop approached but not hit",
     "Price came within 5 points of the stop and the trade survived. A run of "
     "these is not skill."),
    ("Duration", "Time in the market",
     "Reported in seconds under a minute, minutes and seconds under an hour, "
     "and hours and minutes above that - the units the trade was actually "
     "experienced in."),
    ("Session", "Asia, London, NY Overlap, NY Late",
     "Assigned from the entry time in broker server time. The boundaries are "
     "editable in the app's settings."),
    ("Excursion source", "ticks or bars",
     "Whether MAE and MFE were measured from the actual tick stream or from M1 "
     "bar highs and lows. Ticks are exact; bars are close, and are used when "
     "the tick history for that period is not available."),
    ("Monte Carlo", "Resampled forward projection",
     "The trader's own results, drawn with replacement, in thousands of "
     "different orders. It answers what this strategy could do next, including "
     "how bad an unlucky sequence looks - without assuming the results follow "
     "a normal distribution, which they never do."),
]


def build_glossary(book: Workbook) -> None:
    sheet = book.create_sheet("Glossary")
    _banner(
        sheet, "Glossary",
        "Every measure in this workbook, in plain terms, with how to read it.",
        width=8,
    )
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 30
    for column in "CDEFGH":
        sheet.column_dimensions[column].width = 14
    for index, title in enumerate(("Measure", "Definition", "How to read it"),
                                  start=1):
        _write(sheet, 4, index, title, "ps_header")
    sheet.merge_cells(start_row=4, start_column=3, end_row=4, end_column=8)
    sheet.row_dimensions[4].height = 24

    for offset, (term, definition, explanation) in enumerate(GLOSSARY):
        row = 5 + offset
        _write(sheet, row, 1, term, "ps_cell_left",
               font=Font(name="Segoe UI Semibold", size=9, color=INK))
        _write(sheet, row, 2, definition, "ps_cell_left")
        _write(sheet, row, 3, explanation, "ps_note")
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        sheet.row_dimensions[row].height = 42
        for column in range(1, 9):
            sheet.cell(row=row, column=column).border = Border(
                bottom=Side(style="thin", color=LINE)
            )
    sheet.freeze_panes = "A5"


# -- entry point -----------------------------------------------------------
def _shot_map(account_id: int, positions: list[dict]) -> dict[int, list[dict]]:
    return {p["id"]: db.shots_for(p["id"]) for p in positions}


def _folder_map(positions: list[dict], account: dict, conf: dict) -> dict[int, str]:
    """Folder path per position, as text for the workbook.

    Imported here rather than at module scope: capture imports render, which
    pulls in PIL, and generating a workbook should not require it.
    """
    from . import capture

    root = paths.charts_dir()
    out: dict[int, str] = {}
    for position in positions:
        try:
            folder = capture.position_folder(position, account, conf)
            out[position["id"]] = folder.relative_to(root).as_posix()
        except Exception:
            out[position["id"]] = ""
    return out


def build_workbook(account_id: int, destination: Path | str | None = None,
                   conf: dict | None = None,
                   progress=None) -> dict:
    """Build the mentorship workbook for one account.

    Reads only from the database - no terminal connection needed, so a report
    can be regenerated with MT5 closed.
    """
    conf = conf or settings.load()
    account = db.get_account(account_id)
    if not account:
        raise ValueError(f"no such account: {account_id}")

    def step(message: str) -> None:
        if progress:
            progress({"message": message})

    step("Reading positions")
    positions = [p for p in db.position_rows(account_id) if p.get("close_time")]
    positions.sort(key=lambda p: p["close_time"])
    start_balance = account.get("start_balance")
    # Running balance, drawdown and streak are sequence-dependent, so they are
    # recomputed here rather than trusted from whenever ingest last ran.
    metrics.enrich_series(positions, start_balance, conf)

    step("Computing statistics")
    summary = stats.summarize(positions, start_balance)
    analysis = conf.get("analysis", {})
    monte = stats.monte_carlo(
        positions,
        runs=int(analysis.get("monte_carlo_runs", 5000)),
        horizon=int(analysis.get("monte_carlo_horizon", 100)),
        start_balance=start_balance,
    )

    shots = _shot_map(account_id, positions)
    folders = _folder_map(positions, account, conf)

    book = Workbook()
    book.remove(book.active)
    register_styles(book)

    step("Writing trades")
    _ref, first_row = build_trades(book, positions, shots, folders)
    last_row = first_row + max(len(positions) - 1, 0)

    step("Building dashboard")
    build_dashboard(book, positions, summary, account, monte, first_row, last_row)
    step("Building pivots")
    build_pivots(book, positions, first_row, last_row)
    step("Building excursions")
    build_excursions(book, positions, summary, first_row, last_row)
    build_risk(book, positions, summary, first_row, last_row)
    step("Writing findings")
    build_findings(book, positions, summary)
    if conf.get("export", {}).get("include_charts_sheet", True):
        build_charts_index(book, positions, shots, folders)
    build_glossary(book)

    _print_setup(book, account)

    book.properties.title = f"{paths.APP_TITLE} — account {account.get('login')}"
    book.properties.creator = paths.APP_TITLE
    book.calculation.fullCalcOnLoad = True
    book.active = 0

    if destination is None:
        stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        destination = paths.exports_dir() / (
            f"{account.get('login', 'account')}_mentorship_{stamp}.xlsx"
        )
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    step("Saving workbook")
    book.save(destination)

    return {
        "path": str(destination),
        "trades": len(positions),
        "shots": sum(len(v) for v in shots.values()),
        "sheets": book.sheetnames,
        "findings": len(_findings(positions, summary))
        + len(_group_findings(positions, summary)),
    }
